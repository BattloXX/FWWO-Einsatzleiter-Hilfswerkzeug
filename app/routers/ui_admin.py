"""Admin-UI: Stammdaten, User, Rollen, API-Keys."""
import logging
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.audit import write_audit
from app.core.permissions import has_role, require_role, same_org_or_system_admin
from app.core.security import generate_api_key, hash_api_key, hash_password
from app.core.templating import templates
from app.db import get_db
from app.models.master import (
    BOS_VALUES,
    AlarmDispatchVehicle,
    AlarmType,
    DefaultMessage,
    DefaultMessageAlarm,
    FireDept,
    LageHint,
    Member,
    MemberQualification,
    MessageSuggestion,
    MessageSuggestionAlarm,
    Qualification,
    SystemSettings,
    TaskSuggestion,
    TaskSuggestionAlarm,
    VehicleMaster,
)
from app.models.user import ApiKey, AuditLog, DeviceToken, PushLog, PushSubscription, Role, User, UserRole

router = APIRouter(prefix="/admin")
logger_admin = logging.getLogger("einsatzleiter.admin")


def _org_filter(q, user, col):
    """Apply org_id filter for non-sysadmin users."""
    if not has_role(user, "system_admin") and user.org_id:
        q = q.filter(col == user.org_id)
    return q


def _admin_check(request: Request):
    user = getattr(request.state, "user", None)
    if not user:
        raise RedirectResponse("/login", status_code=302)
    return user


# ── Benutzer ──────────────────────────────────────────────────────────────────

@router.get("/benutzer", response_class=HTMLResponse)
async def users_list(request: Request, db: Session = Depends(get_db),
                     _=Depends(require_role("admin"))):
    user = request.state.user
    is_sysadmin = has_role(user, "system_admin")
    users = _org_filter(db.query(User), user, User.org_id).filter(User.is_device == False).order_by(User.username).all()  # noqa: E712
    roles = db.query(Role).all()
    all_orgs = db.query(FireDept).order_by(FireDept.name).all() if is_sysadmin else []
    return templates.TemplateResponse(request, "admin/users.html", {
        "user": user,
        "users": users, "roles": roles,
        "is_sysadmin": is_sysadmin,
        "all_orgs": all_orgs,
        "saved": request.query_params.get("saved"),
        "mail_sent": request.query_params.get("mail"),
        "error": request.query_params.get("error"),
    })


@router.post("/benutzer/neu")
async def create_user(
    request: Request,
    username: str = Form(...), display_name: str = Form(""),
    full_name: str = Form(""), email: str = Form(""), phone: str = Form(""),
    password: str = Form(...), role_codes: list[str] = Form([]),
    org_id: int | None = Form(None),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    current_user = request.state.user
    email_clean = (email or "").strip().lower() or None
    if email_clean:
        existing = db.query(User).filter(User.email == email_clean).first()
        if existing:
            return RedirectResponse(
                "/admin/benutzer?error=email_exists", status_code=303,
            )
    target_org_id = (org_id if has_role(current_user, "system_admin") and org_id and org_id != 0 else current_user.org_id)
    new_user = User(
        username=username,
        display_name=display_name or username,
        full_name=(full_name.strip() or None),
        email=email_clean,
        phone=(phone.strip() or None),
        password_hash=hash_password(password),
        org_id=target_org_id,
    )
    db.add(new_user)
    db.flush()
    for code in role_codes:
        role = db.query(Role).filter(Role.code == code).first()
        if role:
            db.add(UserRole(user_id=new_user.id, role_id=role.id))
    write_audit(db, "admin.user.created", user_id=request.state.user.id,
                entity_type="user", entity_id=new_user.id,
                payload={"role_codes": role_codes})
    db.commit()
    return RedirectResponse("/admin/benutzer?saved=1", status_code=303)


@router.post("/benutzer/{user_id}/loeschen")
async def delete_user(
    user_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    u = db.get(User, user_id)
    if u and not same_org_or_system_admin(request.state.user, u.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    if u and u.id != request.state.user.id:
        u.active = False
        write_audit(db, "admin.user.deactivated", user_id=request.state.user.id,
                    entity_type="user", entity_id=user_id)
        db.commit()
    return RedirectResponse("/admin/benutzer", status_code=303)


# ── API-Keys ──────────────────────────────────────────────────────────────────

@router.get("/api-keys", response_class=HTMLResponse)
async def api_keys(request: Request, db: Session = Depends(get_db),
                   _=Depends(require_role("admin"))):
    keys = _org_filter(db.query(ApiKey), request.state.user, ApiKey.org_id).order_by(ApiKey.created_at.desc()).all()
    return templates.TemplateResponse(request, "admin/api_keys.html", {
        "user": request.state.user, "keys": keys, "new_key": None,
    })


@router.post("/api-keys/neu")
async def create_api_key(
    request: Request, label: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    raw = generate_api_key()
    user = request.state.user
    key = ApiKey(
        key_hash=hash_api_key(raw),
        label=label,
        org_id=user.org_id,
        created_by_user_id=user.id,
    )
    db.add(key)
    write_audit(db, "admin.api_key.created", user_id=user.id,
                payload={"label": label, "org_id": user.org_id})
    db.commit()
    keys = _org_filter(db.query(ApiKey), user, ApiKey.org_id).order_by(ApiKey.created_at.desc()).all()
    return templates.TemplateResponse(request, "admin/api_keys.html", {
        "user": request.state.user, "keys": keys, "new_key": raw,
    })


@router.post("/api-keys/{key_id}/sperren")
async def revoke_api_key(
    key_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    key = db.get(ApiKey, key_id)
    if key and not same_org_or_system_admin(request.state.user, key.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    if key:
        key.revoked_at = datetime.now(UTC)
        write_audit(db, "admin.api_key.revoked", user_id=request.state.user.id,
                    entity_type="api_key", entity_id=key_id)
        db.commit()
    return RedirectResponse("/admin/api-keys", status_code=303)


# ── Mitglieder ────────────────────────────────────────────────────────────────

@router.get("/mitglieder", response_class=HTMLResponse)
async def members_list(request: Request, db: Session = Depends(get_db),
                       _=Depends(require_role("admin"))):
    members = _org_filter(db.query(Member), request.state.user, Member.org_id).order_by(Member.lastname, Member.firstname).all()
    qualifications = db.query(Qualification).all()
    return templates.TemplateResponse(request, "admin/members.html", {
        "user": request.state.user,
        "members": members, "qualifications": qualifications,
    })


@router.post("/mitglieder/neu")
async def create_member(
    request: Request,
    lastname: str = Form(...), firstname: str = Form(...),
    phone: str = Form(""), email: str = Form(""),
    qualification_codes: list[str] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    member = Member(lastname=lastname, firstname=firstname,
                    phone=phone or None, email=email or None)
    db.add(member)
    db.flush()
    for code in qualification_codes:
        q = db.query(Qualification).filter(Qualification.code == code).first()
        if q:
            db.add(MemberQualification(member_id=member.id, qualification_id=q.id))
    write_audit(db, "admin.member.created", user_id=request.state.user.id,
                entity_type="member", entity_id=member.id)
    db.commit()
    return RedirectResponse("/admin/mitglieder", status_code=303)


# ── Audit-Log ─────────────────────────────────────────────────────────────────

@router.get("/audit", response_class=HTMLResponse)
async def audit_log(request: Request, db: Session = Depends(get_db),
                    _=Depends(require_role("admin"))):
    user = request.state.user
    if has_role(user, "system_admin"):
        entries = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(500).all()
    else:
        org_user_ids = db.query(User.id).filter(User.org_id == user.org_id).subquery()
        entries = (
            db.query(AuditLog)
            .filter(AuditLog.user_id.in_(org_user_ids))
            .order_by(AuditLog.created_at.desc())
            .limit(500)
            .all()
        )
    return templates.TemplateResponse(request, "admin/audit.html", {
        "user": request.state.user, "entries": entries,
    })


# ── Admin Landing Page ────────────────────────────────────────────────────────

@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
async def admin_index(request: Request, db: Session = Depends(get_db),
                      _=Depends(require_role("admin", "org_admin", "system_admin"))):
    from app.core.permissions import has_role
    from app.models.incident import Incident
    user = request.state.user
    is_sysadmin = has_role(user, "system_admin")

    # KPI-Kennzahlen (Org-scoped, system_admin sieht alles)
    incident_q = db.query(Incident)
    member_q = db.query(Member)
    vehicle_q = db.query(VehicleMaster).filter(VehicleMaster.active == True)  # noqa: E712
    user_q = db.query(User).filter(User.active == True)  # noqa: E712
    if not is_sysadmin and user.org_id:
        incident_q = incident_q.filter(Incident.primary_org_id == user.org_id)
        member_q = member_q.filter(Member.org_id == user.org_id)
        vehicle_q = vehicle_q.filter(VehicleMaster.dept_id == user.org_id)
        user_q = user_q.filter(User.org_id == user.org_id)

    kpis = {
        "active_incidents": incident_q.filter(Incident.status == "active").count(),
        "total_incidents": incident_q.count(),
        "members": member_q.filter(Member.active == True).count(),  # noqa: E712
        "vehicles": vehicle_q.count(),
        "users": user_q.count(),
    }
    recent_incidents = incident_q.order_by(Incident.started_at.desc()).limit(5).all()
    recent_audit = (
        db.query(AuditLog)
        .order_by(AuditLog.created_at.desc())
        .limit(8)
        .all()
    )

    return templates.TemplateResponse(request, "admin/index.html", {
        "user": user,
        "is_sysadmin": is_sysadmin,
        "kpis": kpis,
        "recent_incidents": recent_incidents,
        "recent_audit": recent_audit,
    })


# ── Mitglieder Edit / Toggle / Qualifikationen ────────────────────────────────

@router.post("/mitglieder/{member_id}/edit")
async def edit_member(
    member_id: int, request: Request,
    lastname: str = Form(...), firstname: str = Form(...),
    phone: str = Form(""), email: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    m = db.get(Member, member_id)
    if m and not same_org_or_system_admin(request.state.user, m.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    if m:
        m.lastname = lastname
        m.firstname = firstname
        m.phone = phone or None
        m.email = email or None
        write_audit(db, "admin.member.edited", user_id=request.state.user.id,
                    entity_type="member", entity_id=member_id)
        db.commit()
    return RedirectResponse("/admin/mitglieder?saved=1", status_code=303)


@router.post("/mitglieder/{member_id}/toggle")
async def toggle_member(
    member_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    m = db.get(Member, member_id)
    if m and not same_org_or_system_admin(request.state.user, m.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    if m:
        m.active = not m.active
        write_audit(db, "admin.member.toggled", user_id=request.state.user.id,
                    entity_type="member", entity_id=member_id)
        db.commit()
    return RedirectResponse("/admin/mitglieder", status_code=303)


@router.post("/mitglieder/bulk-delete")
async def bulk_delete_members(
    request: Request,
    member_ids: list[int] = Form(...),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    import logging
    import urllib.parse
    from app.models.incident import Incident as _Incident, IncidentVehicle as _IV
    from app.models.breathing import TroopMember as _TM, PressureLog as _PL

    logger = logging.getLogger(__name__)
    user = request.state.user
    rows = _org_filter(
        db.query(Member).filter(Member.id.in_(member_ids)),
        user, Member.org_id,
    ).all()
    deleted = 0
    blocked: list[str] = []
    for m in rows:
        name = m.full_name
        mid = m.id
        sp = db.begin_nested()
        try:
            # FK-Nullungen und Delete zusammen im Savepoint –
            # bei Fehlschlag werden alle Änderungen für dieses Mitglied zurückgerollt.
            db.query(_IV).filter(_IV.commander_member_id == mid).update(
                {"commander_member_id": None}, synchronize_session=False
            )
            db.query(_Incident).filter(_Incident.incident_leader_member_id == mid).update(
                {"incident_leader_member_id": None}, synchronize_session=False
            )
            db.query(_TM).filter(_TM.member_id == mid).update(
                {"member_id": None}, synchronize_session=False
            )
            db.query(_PL).filter(_PL.member_id == mid).update(
                {"member_id": None}, synchronize_session=False
            )
            db.delete(m)
            db.flush()
            sp.commit()
            deleted += 1
        except Exception as exc:
            logger.warning("bulk_delete: Mitglied %s (id=%s) nicht löschbar: %s", name, mid, exc)
            sp.rollback()
            db.expunge_all()
            blocked.append(name)
    write_audit(db, "admin.member.bulk_delete", user_id=user.id,
                payload={"deleted": deleted, "blocked": blocked})
    db.commit()
    blocked_q = ("&blocked=" + urllib.parse.quote(", ".join(blocked))) if blocked else ""
    return RedirectResponse(
        f"/admin/mitglieder?saved=1&deleted={deleted}{blocked_q}",
        status_code=303,
    )


@router.post("/mitglieder/{member_id}/quali")
async def update_member_quali(
    member_id: int, request: Request,
    qualification_codes: list[str] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    m = db.get(Member, member_id)
    if not m:
        return RedirectResponse("/admin/mitglieder", status_code=303)
    if not same_org_or_system_admin(request.state.user, m.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    # Remove all existing, then re-add
    db.query(MemberQualification).filter(MemberQualification.member_id == member_id).delete()
    form_data = await request.form()
    for code in qualification_codes:
        q = db.query(Qualification).filter(Qualification.code == code).first()
        if q:
            valid_until_str = form_data.get(f"valid_until_{code}")
            valid_until = None
            if valid_until_str:
                from datetime import date
                try:
                    valid_until = date.fromisoformat(valid_until_str)
                except ValueError:
                    pass
            db.add(MemberQualification(member_id=member_id, qualification_id=q.id, valid_until=valid_until))
    write_audit(db, "admin.member.quali_updated", user_id=request.state.user.id,
                entity_type="member", entity_id=member_id)
    db.commit()
    return RedirectResponse("/admin/mitglieder?saved=1", status_code=303)


# ── Mitglieder Excel-Import ───────────────────────────────────────────────────

@router.get("/mitglieder/excel-import")
def excel_import_redirect():
    """GET-Fallback: User hat Direkt-URL aufgerufen → zur Mitglieder-Seite."""
    return RedirectResponse("/admin/mitglieder", status_code=303)


@router.post("/mitglieder/excel-import")
async def import_members_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Massimport von Mitgliedern aus einer Excel-Datei.

    Erwartete Spalten (Groß-/Kleinschreibung egal):
      Nachname / Lastname, Vorname / Firstname,
      Telefon / Phone (optional), E-Mail / Email (optional)
    """
    import io as _io
    try:
        import openpyxl
    except ImportError:
        return RedirectResponse("/admin/mitglieder?error=openpyxl_missing", status_code=303)

    raw = await file.read()
    if not raw:
        return RedirectResponse("/admin/mitglieder?error=empty_file", status_code=303)

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/mitglieder?error=invalid_excel", status_code=303)

    _LASTNAME_ALIASES  = {"nachname", "lastname", "name", "zuname", "familienname"}
    _FIRSTNAME_ALIASES = {"vorname", "firstname", "rufname"}
    _PHONE_ALIASES     = {"telefon", "phone", "tel", "mobil", "handy",
                          "mobil nummer 1", "mobiltelefon", "telefonnummer"}
    _EMAIL_ALIASES     = {"e-mail", "email", "mail", "e-mail 1", "e mail"}
    _ROLE_ALIASES      = {"bezeichnung", "funktion", "rolle"}
    _ROLE_TO_QUALI = [
        ("gruppenkommandant", "GK"),
        ("zugskommandant",    "ZK"),
        ("kommandant",        "ZK"),
        ("atemschutz",        "AGT"),
        ("maschinist",        "MA"),
        ("truppführer",       "TF"),
        ("truppfuhrer",       "TF"),
        ("truppmann",         "TM"),
        ("einsatzleiter",     "EL"),
        ("jugend",            "JF"),
    ]

    # Build column index from header row (safe against empty sheets)
    try:
        header_row = next(ws.iter_rows(max_row=1))
    except StopIteration:
        return RedirectResponse("/admin/mitglieder?error=empty_sheet", status_code=303)
    headers = [str(c.value or "").strip().lower() for c in header_row]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in _LASTNAME_ALIASES:
            col_map.setdefault("lastname", i)
        elif h in _FIRSTNAME_ALIASES:
            col_map.setdefault("firstname", i)
        elif h in _PHONE_ALIASES:
            col_map.setdefault("phone", i)
        elif h in _EMAIL_ALIASES:
            col_map.setdefault("email", i)
        elif h in _ROLE_ALIASES:
            col_map.setdefault("role", i)

    if "lastname" not in col_map or "firstname" not in col_map:
        found = ", ".join('"' + h + '"' for h in headers[:8] if h)
        detail = "Gefundene Spalten: " + found + ". Erwartet: Zuname/Nachname und Vorname."
        import urllib.parse
        return RedirectResponse(
            f"/admin/mitglieder?error=missing_columns&error_detail={urllib.parse.quote(detail)}",
            status_code=303,
        )

    # Pre-load qualifications for role mapping
    quali_cache: dict[str, int] = {}
    for _label, _code in _ROLE_TO_QUALI:
        if _code not in quali_cache:
            q = db.query(Qualification).filter(Qualification.code == _code).first()
            if q:
                quali_cache[_code] = q.id

    user = request.state.user
    created = 0
    updated = 0
    skipped = 0
    row_errors = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        sp = db.begin_nested()
        try:
            lastname = str(row[col_map["lastname"]] or "").strip()
            firstname = str(row[col_map["firstname"]] or "").strip()
            if not lastname or not firstname:
                skipped += 1
                sp.commit()
                continue
            phone = str(row[col_map["phone"]] if "phone" in col_map and col_map["phone"] < len(row) and row[col_map["phone"]] else "").strip() or None
            email = str(row[col_map["email"]] if "email" in col_map and col_map["email"] < len(row) and row[col_map["email"]] else "").strip().lower() or None
            role_raw = str(row[col_map["role"]] if "role" in col_map and col_map["role"] < len(row) and row[col_map["role"]] else "").strip().lower()
            # Match existing (same name in same org) → update; else create
            existing = db.query(Member).filter(
                Member.org_id == user.org_id,
                Member.lastname == lastname,
                Member.firstname == firstname,
            ).first()
            if existing:
                if not existing.phone and phone:
                    existing.phone = phone
                if not existing.email and email:
                    existing.email = email
                existing.active = True
                member = existing
                updated += 1
            else:
                member = Member(lastname=lastname, firstname=firstname, phone=phone, email=email,
                                org_id=user.org_id, active=True)
                db.add(member)
                db.flush()  # get member.id
                created += 1
            # Auto-assign qualification from role column (additiv, idempotent)
            if role_raw and quali_cache:
                for _label, _code in _ROLE_TO_QUALI:
                    if _label in role_raw and _code in quali_cache:
                        already = db.query(MemberQualification).filter_by(
                            member_id=member.id, qualification_id=quali_cache[_code]
                        ).first()
                        if not already:
                            db.add(MemberQualification(member_id=member.id, qualification_id=quali_cache[_code]))
                        break
            sp.commit()
        except Exception:
            sp.rollback()
            row_errors += 1

    if created or updated:
        db.commit()
        write_audit(db, "admin.member.excel_import", user_id=user.id,
                    payload={"created": created, "updated": updated,
                             "skipped": skipped, "row_errors": row_errors})
    return RedirectResponse(
        f"/admin/mitglieder?saved=1&imported={created}&updated={updated}"
        f"&skipped={skipped}&row_errors={row_errors}",
        status_code=303,
    )


# ── Benutzer Edit / Rollen / Passwort-Reset ───────────────────────────────────

@router.post("/benutzer/{user_id}/edit")
async def edit_user(
    user_id: int, request: Request,
    display_name: str = Form(...),
    full_name: str = Form(""), email: str = Form(""), phone: str = Form(""),
    org_id: int | None = Form(None),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    current_user = request.state.user
    u = db.get(User, user_id)
    if not u:
        return RedirectResponse("/admin/benutzer", status_code=303)
    if not same_org_or_system_admin(current_user, u.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    email_clean = (email or "").strip().lower() or None
    if email_clean and email_clean != u.email:
        existing = db.query(User).filter(User.email == email_clean, User.id != user_id).first()
        if existing:
            return RedirectResponse("/admin/benutzer?error=email_exists", status_code=303)
    u.display_name = display_name
    u.full_name = full_name.strip() or None
    u.email = email_clean
    u.phone = phone.strip() or None
    if has_role(current_user, "system_admin") and org_id is not None:
        u.org_id = org_id if org_id != 0 else None
    write_audit(db, "admin.user.edited", user_id=current_user.id,
                entity_type="user", entity_id=user_id)
    db.commit()
    return RedirectResponse("/admin/benutzer?saved=1", status_code=303)


@router.post("/benutzer/{user_id}/reset-mail")
async def send_user_reset_mail(
    user_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    """Admin löst den Self-Service-Reset-Flow für einen anderen Benutzer aus."""
    import hashlib
    import secrets as sec
    from datetime import datetime, timedelta

    from app.config import settings
    from app.models.password_reset import PasswordResetToken
    from app.services.mail_service import send_password_reset

    u = db.get(User, user_id)
    if not u or not u.email:
        return RedirectResponse("/admin/benutzer?error=no_email", status_code=303)
    if not same_org_or_system_admin(request.state.user, u.org_id):
        raise HTTPException(403, "Keine Berechtigung")

    # Alte Tokens entwerten
    now = datetime.now(UTC)
    db.query(PasswordResetToken).filter(
        PasswordResetToken.user_id == u.id,
        PasswordResetToken.used_at.is_(None),
    ).update({"used_at": now})

    raw_token = sec.token_urlsafe(32)
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    db.add(PasswordResetToken(
        user_id=u.id,
        token_hash=token_hash,
        expires_at=now + timedelta(minutes=settings.PASSWORD_RESET_TTL_MIN),
        requesting_ip=request.client.host if request.client else None,
    ))
    write_audit(db, "admin.user.password_reset_mail",
                user_id=request.state.user.id, entity_type="user", entity_id=u.id)
    db.commit()

    base = settings.effective_public_base_url.rstrip("/")
    reset_url = f"{base}/passwort-zuruecksetzen?token={raw_token}"
    try:
        await send_password_reset(
            to=u.email, reset_url=reset_url,
            user_display_name=u.full_name or u.display_name or u.username,
            db=db,
        )
    except Exception:
        return RedirectResponse("/admin/benutzer?error=mail_failed", status_code=303)
    return RedirectResponse("/admin/benutzer?saved=1&mail=1", status_code=303)


@router.post("/benutzer/{user_id}/rollen")
async def update_user_roles(
    user_id: int, request: Request,
    role_codes: list[str] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    u = db.get(User, user_id)
    if not u:
        return RedirectResponse("/admin/benutzer", status_code=303)
    if not same_org_or_system_admin(request.state.user, u.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    db.query(UserRole).filter(UserRole.user_id == user_id).delete()
    for code in role_codes:
        role = db.query(Role).filter(Role.code == code).first()
        if role:
            db.add(UserRole(user_id=user_id, role_id=role.id))
    write_audit(db, "admin.user.roles_updated", user_id=request.state.user.id,
                entity_type="user", entity_id=user_id)
    db.commit()
    return RedirectResponse("/admin/benutzer?saved=1", status_code=303)


@router.post("/benutzer/{user_id}/passwort")
async def reset_user_password(
    user_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    import secrets as sec
    u = db.get(User, user_id)
    if not u:
        return RedirectResponse("/admin/benutzer", status_code=303)
    if not same_org_or_system_admin(request.state.user, u.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    new_pw = sec.token_urlsafe(12)
    u.password_hash = hash_password(new_pw)
    write_audit(db, "admin.user.password_reset", user_id=request.state.user.id,
                entity_type="user", entity_id=user_id)
    db.commit()
    users = _org_filter(db.query(User), request.state.user, User.org_id).order_by(User.username).all()
    roles = db.query(Role).all()
    return templates.TemplateResponse(request, "admin/users.html", {
        "user": request.state.user,
        "users": users, "roles": roles,
        "new_password": new_pw, "new_password_user": u.username,
    })


# ── Fahrzeuge CRUD ────────────────────────────────────────────────────────────

@router.get("/einheiten", response_class=HTMLResponse, include_in_schema=False)
async def vehicles_list_alias(request: Request, db: Session = Depends(get_db),
                              _=Depends(require_role("admin", "org_admin"))):
    return RedirectResponse("/admin/fahrzeuge", status_code=301)


@router.get("/fahrzeuge", response_class=HTMLResponse)
async def vehicles_list(request: Request, db: Session = Depends(get_db),
                        _=Depends(require_role("admin", "org_admin"))):
    from app.core.permissions import has_role
    user = request.state.user
    is_sysadmin = has_role(user, "system_admin")
    all_orgs = db.query(FireDept).order_by(FireDept.name).all() if is_sysadmin else []
    vehicles = (
        db.query(VehicleMaster)
        .join(VehicleMaster.dept)
        .order_by(FireDept.name, VehicleMaster.display_order)
        .all()
    )
    saved = request.query_params.get("saved")
    error = request.query_params.get("error")
    return templates.TemplateResponse(request, "admin/vehicles.html", {
        "user": user, "vehicles": vehicles, "all_orgs": all_orgs,
        "is_sysadmin": is_sysadmin, "saved": saved, "error": error,
        "bos_values": BOS_VALUES,
    })


@router.post("/fahrzeuge/neu")
async def create_vehicle(
    request: Request,
    code: str = Form(...), name: str = Form(...), type: str = Form(""),
    is_first_train: str = Form(""),
    bos_override: str = Form(""),
    dept_id: int | None = Form(None),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    from app.core.permissions import has_role
    user = request.state.user
    target_dept_id = dept_id if (has_role(user, "system_admin") and dept_id) else None
    if not target_dept_id:
        home = db.query(FireDept).filter(FireDept.is_home_org == True).first()  # noqa: E712
        target_dept_id = home.id if home else None
    if not target_dept_id:
        return RedirectResponse("/admin/fahrzeuge?error=no_org", status_code=303)
    max_order = db.query(VehicleMaster).filter(VehicleMaster.dept_id == target_dept_id).count()
    v = VehicleMaster(
        dept_id=target_dept_id, code=code, name=name, type=type,
        is_first_train=bool(is_first_train),
        bos_override=bos_override or None,
        display_order=max_order,
    )
    db.add(v)
    write_audit(db, "admin.vehicle.created", user_id=user.id,
                entity_type="vehicle_master", entity_id=v.id if v.id else 0)
    db.commit()
    return RedirectResponse("/admin/fahrzeuge?saved=1", status_code=303)


@router.post("/fahrzeuge/{vehicle_id}/edit")
async def edit_vehicle(
    vehicle_id: int, request: Request,
    code: str = Form(...), name: str = Form(...), type: str = Form(""),
    is_first_train: str = Form(""),
    bos_override: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    v = db.get(VehicleMaster, vehicle_id)
    if v:
        v.code = code
        v.name = name
        v.type = type
        v.is_first_train = bool(is_first_train)
        v.bos_override = bos_override or None
        write_audit(db, "admin.vehicle.edited", user_id=request.state.user.id,
                    entity_type="vehicle_master", entity_id=vehicle_id)
        db.commit()
    return RedirectResponse("/admin/fahrzeuge?saved=1", status_code=303)


@router.post("/fahrzeuge/{vehicle_id}/toggle")
async def toggle_vehicle(
    vehicle_id: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    v = db.get(VehicleMaster, vehicle_id)
    if v:
        v.active = not v.active
        db.commit()
    return RedirectResponse("/admin/fahrzeuge", status_code=303)


@router.post("/fahrzeuge/{vehicle_id}/order/{direction}")
async def reorder_vehicle(
    vehicle_id: int, direction: str, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    v = db.get(VehicleMaster, vehicle_id)
    if not v:
        return RedirectResponse("/admin/fahrzeuge", status_code=303)
    siblings = (
        db.query(VehicleMaster)
        .filter(VehicleMaster.dept_id == v.dept_id)
        .order_by(VehicleMaster.display_order)
        .all()
    )
    idx = next((i for i, s in enumerate(siblings) if s.id == vehicle_id), None)
    if idx is not None:
        swap_idx = idx - 1 if direction == "up" else idx + 1
        if 0 <= swap_idx < len(siblings):
            siblings[idx].display_order, siblings[swap_idx].display_order = (
                siblings[swap_idx].display_order, siblings[idx].display_order
            )
            db.commit()
    return RedirectResponse("/admin/fahrzeuge", status_code=303)


# ── Auftragsvorlagen CRUD ─────────────────────────────────────────────────────

@router.get("/auftragsvorlagen", response_class=HTMLResponse)
async def task_suggestions_list(request: Request, db: Session = Depends(get_db),
                                _=Depends(require_role("admin", "org_admin"))):
    from sqlalchemy.orm import joinedload
    alarm_types = db.query(AlarmType).order_by(AlarmType.code).all()
    suggestions = (
        db.query(TaskSuggestion)
        .options(joinedload(TaskSuggestion.alarm_assignments))
        .order_by(TaskSuggestion.id)
        .all()
    )
    assignments_by_alarm: dict[str, list] = {at.code: [] for at in alarm_types}
    for s in suggestions:
        for a in s.alarm_assignments:
            if a.alarm_type_code in assignments_by_alarm:
                assignments_by_alarm[a.alarm_type_code].append(a)
    for code in assignments_by_alarm:
        assignments_by_alarm[code].sort(key=lambda a: a.display_order)

    edit_id = request.query_params.get("edit")
    edit_suggestion = None
    edit_assigned_codes: list[str] = []
    if edit_id:
        try:
            edit_suggestion = db.get(TaskSuggestion, int(edit_id))
        except (ValueError, TypeError):
            pass
        if edit_suggestion:
            edit_assigned_codes = [
                a.alarm_type_code
                for a in sorted(edit_suggestion.alarm_assignments, key=lambda x: x.display_order)
            ]
    return templates.TemplateResponse(request, "admin/task_suggestions.html", {
        "user": request.state.user, "alarm_types": alarm_types,
        "suggestions": suggestions, "assignments_by_alarm": assignments_by_alarm,
        "edit_suggestion": edit_suggestion, "edit_assigned_codes": edit_assigned_codes,
        "saved": request.query_params.get("saved"),
        "error": request.query_params.get("error"),
    })


@router.post("/auftragsvorlagen/neu")
async def create_task_suggestion(
    request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    existing = db.query(TaskSuggestion).filter(TaskSuggestion.text == text).first()
    if existing:
        return RedirectResponse("/admin/auftragsvorlagen?error=duplicate", status_code=303)
    s = TaskSuggestion(text=text)
    db.add(s)
    db.flush()
    db.commit()
    return RedirectResponse(f"/admin/auftragsvorlagen?saved=1&edit={s.id}", status_code=303)


@router.post("/auftragsvorlagen/{sid}/edit")
async def edit_task_suggestion(
    sid: int, request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    s = db.get(TaskSuggestion, sid)
    if s:
        dup = db.query(TaskSuggestion).filter(
            TaskSuggestion.text == text, TaskSuggestion.id != sid
        ).first()
        if dup:
            return RedirectResponse(f"/admin/auftragsvorlagen?error=duplicate&edit={sid}", status_code=303)
        s.text = text
        db.commit()
    return RedirectResponse(f"/admin/auftragsvorlagen?saved=1&edit={sid}", status_code=303)


@router.post("/auftragsvorlagen/{sid}/alarms")
async def save_task_suggestion_alarms(
    sid: int, request: Request,
    alarm_type_codes: list[str] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    s = db.get(TaskSuggestion, sid)
    if not s:
        return RedirectResponse("/admin/auftragsvorlagen", status_code=303)
    db.query(TaskSuggestionAlarm).filter(TaskSuggestionAlarm.task_suggestion_id == sid).delete()
    for i, code in enumerate(alarm_type_codes):
        db.add(TaskSuggestionAlarm(task_suggestion_id=sid, alarm_type_code=code, display_order=i))
    db.commit()
    return RedirectResponse(f"/admin/auftragsvorlagen?saved=1&edit={sid}", status_code=303)


@router.post("/auftragsvorlagen/{sid}/loeschen")
async def delete_task_suggestion(
    sid: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    s = db.get(TaskSuggestion, sid)
    if s:
        db.delete(s)
        db.commit()
    return RedirectResponse("/admin/auftragsvorlagen", status_code=303)


# ── Meldungsvorlagen CRUD ─────────────────────────────────────────────────────

@router.get("/meldungsvorlagen", response_class=HTMLResponse)
async def msg_suggestions_list(request: Request, db: Session = Depends(get_db),
                               _=Depends(require_role("admin", "org_admin"))):
    from sqlalchemy.orm import joinedload
    alarm_types = db.query(AlarmType).order_by(AlarmType.code).all()
    suggestions = (
        db.query(MessageSuggestion)
        .options(joinedload(MessageSuggestion.alarm_assignments))
        .order_by(MessageSuggestion.id)
        .all()
    )
    assignments_by_alarm: dict[str, list] = {at.code: [] for at in alarm_types}
    for s in suggestions:
        for a in s.alarm_assignments:
            if a.alarm_type_code in assignments_by_alarm:
                assignments_by_alarm[a.alarm_type_code].append(a)
    for code in assignments_by_alarm:
        assignments_by_alarm[code].sort(key=lambda a: a.display_order)

    edit_id = request.query_params.get("edit")
    edit_suggestion = None
    edit_assigned_codes: list[str] = []
    if edit_id:
        try:
            edit_suggestion = db.get(MessageSuggestion, int(edit_id))
        except (ValueError, TypeError):
            pass
        if edit_suggestion:
            edit_assigned_codes = [
                a.alarm_type_code
                for a in sorted(edit_suggestion.alarm_assignments, key=lambda x: x.display_order)
            ]
    return templates.TemplateResponse(request, "admin/message_suggestions.html", {
        "user": request.state.user, "alarm_types": alarm_types,
        "suggestions": suggestions, "assignments_by_alarm": assignments_by_alarm,
        "edit_suggestion": edit_suggestion, "edit_assigned_codes": edit_assigned_codes,
        "saved": request.query_params.get("saved"),
        "error": request.query_params.get("error"),
    })


@router.post("/meldungsvorlagen/neu")
async def create_msg_suggestion(
    request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    existing = db.query(MessageSuggestion).filter(MessageSuggestion.text == text).first()
    if existing:
        return RedirectResponse("/admin/meldungsvorlagen?error=duplicate", status_code=303)
    s = MessageSuggestion(text=text)
    db.add(s)
    db.flush()
    db.commit()
    return RedirectResponse(f"/admin/meldungsvorlagen?saved=1&edit={s.id}", status_code=303)


@router.post("/meldungsvorlagen/{sid}/edit")
async def edit_msg_suggestion(
    sid: int, request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    s = db.get(MessageSuggestion, sid)
    if s:
        dup = db.query(MessageSuggestion).filter(
            MessageSuggestion.text == text, MessageSuggestion.id != sid
        ).first()
        if dup:
            return RedirectResponse(f"/admin/meldungsvorlagen?error=duplicate&edit={sid}", status_code=303)
        s.text = text
        db.commit()
    return RedirectResponse(f"/admin/meldungsvorlagen?saved=1&edit={sid}", status_code=303)


@router.post("/meldungsvorlagen/{sid}/alarms")
async def save_msg_suggestion_alarms(
    sid: int, request: Request,
    alarm_type_codes: list[str] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    s = db.get(MessageSuggestion, sid)
    if not s:
        return RedirectResponse("/admin/meldungsvorlagen", status_code=303)
    db.query(MessageSuggestionAlarm).filter(MessageSuggestionAlarm.message_suggestion_id == sid).delete()
    for i, code in enumerate(alarm_type_codes):
        db.add(MessageSuggestionAlarm(message_suggestion_id=sid, alarm_type_code=code, display_order=i))
    db.commit()
    return RedirectResponse(f"/admin/meldungsvorlagen?saved=1&edit={sid}", status_code=303)


@router.post("/meldungsvorlagen/{sid}/loeschen")
async def delete_msg_suggestion(
    sid: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    s = db.get(MessageSuggestion, sid)
    if s:
        db.delete(s)
        db.commit()
    return RedirectResponse("/admin/meldungsvorlagen", status_code=303)


# ── Alarmtypen CRUD ───────────────────────────────────────────────────────────

@router.get("/alarmtypen", response_class=HTMLResponse)
async def alarm_types_list(request: Request, db: Session = Depends(get_db),
                           _=Depends(require_role("admin"))):
    alarm_types = db.query(AlarmType).order_by(AlarmType.code).all()
    saved = request.query_params.get("saved")
    error = request.query_params.get("error")
    return templates.TemplateResponse(request, "admin/alarm_types.html", {
        "user": request.state.user, "alarm_types": alarm_types, "saved": saved, "error": error,
    })


@router.post("/alarmtypen/neu")
async def create_alarm_type(
    request: Request,
    code: str = Form(...), category: str = Form("T"), label: str = Form(""),
    default_first_train_only: str = Form(""), notify_neighbors: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    code = code.upper()
    existing = db.get(AlarmType, code)
    if existing:
        return RedirectResponse("/admin/alarmtypen?error=exists", status_code=303)
    at = AlarmType(
        code=code, category=category, label=label,
        default_first_train_only=bool(default_first_train_only),
        notify_neighbors=bool(notify_neighbors),
    )
    db.add(at)
    db.commit()
    return RedirectResponse("/admin/alarmtypen?saved=1", status_code=303)


@router.post("/alarmtypen/{code}/edit")
async def edit_alarm_type(
    code: str, request: Request,
    category: str = Form("T"), label: str = Form(""),
    default_first_train_only: str = Form(""), notify_neighbors: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    at = db.get(AlarmType, code)
    if at:
        at.category = category
        at.label = label
        at.default_first_train_only = bool(default_first_train_only)
        at.notify_neighbors = bool(notify_neighbors)
        db.commit()
    return RedirectResponse("/admin/alarmtypen?saved=1", status_code=303)


@router.post("/alarmtypen/{code}/loeschen")
async def delete_alarm_type(
    code: str, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    from app.models.incident import Incident
    count = db.query(Incident).filter(Incident.alarm_type_code == code).count()
    if count > 0:
        return RedirectResponse("/admin/alarmtypen?error=in_use", status_code=303)
    at = db.get(AlarmType, code)
    if at:
        db.delete(at)
        db.commit()
    return RedirectResponse("/admin/alarmtypen", status_code=303)


# ── Ausrückordnung ────────────────────────────────────────────────────────────

@router.get("/ausrueckordnung", response_class=HTMLResponse)
async def dispatch_order_list(request: Request, db: Session = Depends(get_db),
                              _=Depends(require_role("admin", "org_admin"))):
    alarm_types = db.query(AlarmType).order_by(AlarmType.code).all()
    home = db.query(FireDept).filter(FireDept.is_home_org == True).first()  # noqa: E712
    vehicles = (
        db.query(VehicleMaster)
        .filter(VehicleMaster.dept_id == home.id if home else True, VehicleMaster.active == True)  # noqa: E712
        .order_by(VehicleMaster.display_order)
        .all()
    )
    dispatch_entries = db.query(AlarmDispatchVehicle).order_by(
        AlarmDispatchVehicle.alarm_type_code, AlarmDispatchVehicle.display_order
    ).all()
    # vehicle_master_id → VehicleMaster Lookup
    vehicle_by_id = {v.id: v for v in vehicles}
    # Matrix: code → geordnete VehicleMaster-Liste (für Übersicht)
    dispatch_matrix: dict = {}
    for e in dispatch_entries:
        dispatch_matrix.setdefault(e.alarm_type_code, []).append(
            vehicle_by_id.get(e.vehicle_master_id)
        )
    # max_count für Matrix-Spalten (mindestens 5)
    max_count = max([len(v) for v in dispatch_matrix.values()] + [5])
    # Edit-Mode
    edit_code = request.query_params.get("edit")
    edit_alarm = None
    edit_order: list[int] = []
    if edit_code:
        edit_alarm = db.get(AlarmType, edit_code)
        edit_order = [e.vehicle_master_id for e in dispatch_entries
                      if e.alarm_type_code == edit_code]
    saved = request.query_params.get("saved")
    return templates.TemplateResponse(request, "admin/dispatch_order.html", {
        "user": request.state.user, "alarm_types": alarm_types, "vehicles": vehicles,
        "dispatch_matrix": dispatch_matrix, "max_count": max_count,
        "edit_alarm": edit_alarm, "edit_order": edit_order, "saved": saved,
    })


@router.post("/ausrueckordnung/{alarm_type_code}")
async def save_dispatch_order(
    alarm_type_code: str, request: Request,
    vehicle_ids: list[int] = Form([]),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    # Reihenfolge wird durch die Form-Reihenfolge der hidden inputs definiert.
    db.query(AlarmDispatchVehicle).filter(
        AlarmDispatchVehicle.alarm_type_code == alarm_type_code
    ).delete()
    for i, vid in enumerate(vehicle_ids):
        db.add(AlarmDispatchVehicle(
            alarm_type_code=alarm_type_code,
            vehicle_master_id=vid,
            display_order=i,
        ))
    db.commit()
    return RedirectResponse("/admin/ausrueckordnung?saved=1", status_code=303)


@router.get("/ausrueckordnung/{alarm_type_code}/reset")
async def reset_dispatch_order(
    alarm_type_code: str, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    db.query(AlarmDispatchVehicle).filter(
        AlarmDispatchVehicle.alarm_type_code == alarm_type_code
    ).delete()
    db.commit()
    return RedirectResponse("/admin/ausrueckordnung?saved=1", status_code=303)


# ── Qualifikationen CRUD ──────────────────────────────────────────────────────

@router.get("/qualifikationen", response_class=HTMLResponse)
async def qualifications_list(request: Request, db: Session = Depends(get_db),
                              _=Depends(require_role("admin"))):
    qualifications = db.query(Qualification).order_by(Qualification.code).all()
    from sqlalchemy import func as sqlfunc
    usage = dict(
        db.query(MemberQualification.qualification_id, sqlfunc.count(MemberQualification.member_id))
        .group_by(MemberQualification.qualification_id)
        .all()
    )
    saved = request.query_params.get("saved")
    return templates.TemplateResponse(request, "admin/qualifications.html", {
        "user": request.state.user, "qualifications": qualifications,
        "usage": usage, "saved": saved,
    })


@router.post("/qualifikationen/neu")
async def create_qualification(
    request: Request, code: str = Form(...), label: str = Form(...),
    is_einsatzleiter: str = Form(""), is_gruppenkommandant: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    q = Qualification(
        code=code.upper(), label=label,
        is_einsatzleiter=bool(is_einsatzleiter),
        is_gruppenkommandant=bool(is_gruppenkommandant),
    )
    db.add(q)
    db.commit()
    return RedirectResponse("/admin/qualifikationen?saved=1", status_code=303)


@router.post("/qualifikationen/{qid}/edit")
async def edit_qualification(
    qid: int, request: Request, code: str = Form(...), label: str = Form(...),
    is_einsatzleiter: str = Form(""), is_gruppenkommandant: str = Form(""),
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    q = db.get(Qualification, qid)
    if q:
        q.code = code.upper()
        q.label = label
        q.is_einsatzleiter = bool(is_einsatzleiter)
        q.is_gruppenkommandant = bool(is_gruppenkommandant)
        db.commit()
    return RedirectResponse("/admin/qualifikationen?saved=1", status_code=303)


@router.post("/qualifikationen/{qid}/loeschen")
async def delete_qualification(
    qid: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    q = db.get(Qualification, qid)
    if q:
        in_use = db.query(MemberQualification).filter(MemberQualification.qualification_id == qid).count()
        if in_use:
            return RedirectResponse(f"/admin/qualifikationen?error=in_use&code={q.code}", status_code=303)
        db.delete(q)
        db.commit()
    return RedirectResponse("/admin/qualifikationen", status_code=303)


# ── Lage-Hinweise CRUD ────────────────────────────────────────────────────────

@router.get("/lage-hinweise", response_class=HTMLResponse)
async def lage_hints_list(request: Request, db: Session = Depends(get_db),
                          _=Depends(require_role("admin", "org_admin"))):
    hints = db.query(LageHint).order_by(LageHint.display_order).all()
    saved = request.query_params.get("saved")
    return templates.TemplateResponse(request, "admin/lage_hints.html", {
        "user": request.state.user, "hints": hints, "saved": saved,
    })


@router.post("/lage-hinweise/neu")
async def create_lage_hint(
    request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    max_order = db.query(LageHint).count()
    db.add(LageHint(text=text, display_order=max_order))
    db.commit()
    return RedirectResponse("/admin/lage-hinweise?saved=1", status_code=303)


@router.post("/lage-hinweise/{hid}/edit")
async def edit_lage_hint(
    hid: int, request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    h = db.get(LageHint, hid)
    if h:
        h.text = text
        db.commit()
    return RedirectResponse("/admin/lage-hinweise?saved=1", status_code=303)


@router.post("/lage-hinweise/{hid}/loeschen")
async def delete_lage_hint(
    hid: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    h = db.get(LageHint, hid)
    if h:
        db.delete(h)
        db.commit()
    return RedirectResponse("/admin/lage-hinweise", status_code=303)


@router.post("/lage-hinweise/{hid}/order/{direction}")
async def reorder_lage_hint(
    hid: int, direction: str, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    h = db.get(LageHint, hid)
    if not h:
        return RedirectResponse("/admin/lage-hinweise", status_code=303)
    siblings = db.query(LageHint).order_by(LageHint.display_order).all()
    idx = next((i for i, x in enumerate(siblings) if x.id == hid), None)
    if idx is not None:
        swap_idx = idx - 1 if direction == "up" else idx + 1
        if 0 <= swap_idx < len(siblings):
            siblings[idx].display_order, siblings[swap_idx].display_order = (
                siblings[swap_idx].display_order, siblings[idx].display_order
            )
            db.commit()
    return RedirectResponse("/admin/lage-hinweise", status_code=303)


# ── Default-Meldungen CRUD ────────────────────────────────────────────────────

@router.get("/default-meldungen", response_class=HTMLResponse)
async def default_messages_list(request: Request, db: Session = Depends(get_db),
                                _=Depends(require_role("admin", "org_admin"))):
    from sqlalchemy.orm import joinedload
    alarm_types = db.query(AlarmType).order_by(AlarmType.code).all()
    messages = (
        db.query(DefaultMessage)
        .options(joinedload(DefaultMessage.alarm_assignments))
        .order_by(DefaultMessage.id)
        .all()
    )
    assignments_by_alarm: dict[str, list] = {at.code: [] for at in alarm_types}
    for m in messages:
        for a in m.alarm_assignments:
            if a.alarm_type_code in assignments_by_alarm:
                assignments_by_alarm[a.alarm_type_code].append(a)
    for code in assignments_by_alarm:
        assignments_by_alarm[code].sort(key=lambda a: a.display_order)

    edit_id = request.query_params.get("edit")
    edit_message = None
    edit_assigned_codes: list[str] = []
    edit_due_by_code: dict[str, int] = {}
    if edit_id:
        try:
            edit_message = db.get(DefaultMessage, int(edit_id))
        except (ValueError, TypeError):
            pass
        if edit_message:
            for a in sorted(edit_message.alarm_assignments, key=lambda x: x.display_order):
                edit_assigned_codes.append(a.alarm_type_code)
                edit_due_by_code[a.alarm_type_code] = a.due_after_sec
    return templates.TemplateResponse(request, "admin/default_messages.html", {
        "user": request.state.user, "alarm_types": alarm_types,
        "messages": messages, "assignments_by_alarm": assignments_by_alarm,
        "edit_message": edit_message, "edit_assigned_codes": edit_assigned_codes,
        "edit_due_by_code": edit_due_by_code,
        "saved": request.query_params.get("saved"),
        "error": request.query_params.get("error"),
    })


@router.post("/default-meldungen/neu")
async def create_default_message(
    request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    existing = db.query(DefaultMessage).filter(DefaultMessage.text == text).first()
    if existing:
        return RedirectResponse("/admin/default-meldungen?error=duplicate", status_code=303)
    m = DefaultMessage(text=text)
    db.add(m)
    db.flush()
    db.commit()
    return RedirectResponse(f"/admin/default-meldungen?saved=1&edit={m.id}", status_code=303)


@router.post("/default-meldungen/{mid}/edit")
async def edit_default_message(
    mid: int, request: Request, text: str = Form(...),
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    text = text.strip()
    m = db.get(DefaultMessage, mid)
    if m:
        dup = db.query(DefaultMessage).filter(
            DefaultMessage.text == text, DefaultMessage.id != mid
        ).first()
        if dup:
            return RedirectResponse(f"/admin/default-meldungen?error=duplicate&edit={mid}", status_code=303)
        m.text = text
        db.commit()
    return RedirectResponse(f"/admin/default-meldungen?saved=1&edit={mid}", status_code=303)


@router.post("/default-meldungen/{mid}/alarms")
async def save_default_message_alarms(
    mid: int, request: Request,
    db: Session = Depends(get_db), _=Depends(require_role("admin", "org_admin")),
):
    form = await request.form()
    alarm_type_codes = form.getlist("alarm_type_codes")
    m = db.get(DefaultMessage, mid)
    if not m:
        return RedirectResponse("/admin/default-meldungen", status_code=303)
    db.query(DefaultMessageAlarm).filter(DefaultMessageAlarm.default_message_id == mid).delete()
    for i, code in enumerate(alarm_type_codes):
        due_min = form.get(f"due_min_{code}", "5")
        try:
            due_after_sec = max(60, int(due_min) * 60)
        except (ValueError, TypeError):
            due_after_sec = 300
        db.add(DefaultMessageAlarm(
            default_message_id=mid, alarm_type_code=code,
            display_order=i, due_after_sec=due_after_sec,
        ))
    db.commit()
    return RedirectResponse(f"/admin/default-meldungen?saved=1&edit={mid}", status_code=303)


@router.post("/default-meldungen/{mid}/loeschen")
async def delete_default_message(
    mid: int, request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin", "org_admin")),
):
    m = db.get(DefaultMessage, mid)
    if m:
        db.delete(m)
        db.commit()
    return RedirectResponse("/admin/default-meldungen", status_code=303)


# ── System-Einstellungen ──────────────────────────────────────────────────────

@router.get("/system-einstellungen", response_class=HTMLResponse)
async def system_settings_page(request: Request, db: Session = Depends(get_db),
                               _=Depends(require_role("system_admin"))):
    settings_raw = db.query(SystemSettings).all()
    settings = {s.key: s.value for s in settings_raw}
    saved = request.query_params.get("saved")
    return templates.TemplateResponse(request, "admin/system_settings.html", {
        "user": request.state.user, "settings": settings, "saved": saved,
    })


@router.post("/system-einstellungen")
async def save_system_settings(
    request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("system_admin")),
    new_key: str = Form(""), new_value: str = Form(""),
):
    form = await request.form()
    known_keys = [
        # Web Push (VAPID)
        "vapid_public_key", "vapid_private_key", "vapid_email", "enable_push",
        # FCM – Native Android Push
        "fcm_enabled", "fcm_project_id", "fcm_credentials_path",
        # E-Mail (SMTP)
        "smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from",
        "smtp_starttls", "smtp_timeout",
        # Allgemein
        "enable_speech_input", "default_session_max_age",
        # Auto-Schließen
        "incident_autoclose_enabled", "incident_autoclose_after_hours",
        "incident_autoclose_grace_minutes",
    ]
    for key in known_keys:
        val = form.get(f"k_{key}")
        if val is not None:
            existing = db.get(SystemSettings, key)
            if existing:
                existing.value = val
                existing.updated_at = datetime.now(UTC)
                existing.updated_by_user_id = request.state.user.id
            else:
                db.add(SystemSettings(key=key, value=val,
                                      updated_by_user_id=request.state.user.id))
    # Handle any existing custom keys
    all_existing = db.query(SystemSettings).all()
    for s in all_existing:
        if s.key not in known_keys:
            val = form.get(f"k_{s.key}")
            if val is not None:
                s.value = val
    # New custom key
    if new_key.strip():
        existing = db.get(SystemSettings, new_key.strip())
        if existing:
            existing.value = new_value
        else:
            db.add(SystemSettings(key=new_key.strip(), value=new_value,
                                  updated_by_user_id=request.state.user.id))
    write_audit(db, "admin.system_settings.updated", user_id=request.state.user.id)
    db.commit()
    return RedirectResponse("/admin/system-einstellungen?saved=1", status_code=303)


@router.post("/system-einstellungen/test-mail", response_class=HTMLResponse)
async def test_smtp_mail(
    request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("system_admin")),
    test_mail_to: str = Form(""),
):
    from app.services.mail_service import MailConfigError, get_smtp_cfg, send_test_mail

    recipient = test_mail_to.strip() or request.state.user.email or ""
    if not recipient:
        return RedirectResponse(
            "/admin/system-einstellungen?mail_error=Keine+Empfängeradresse+angegeben", status_code=303
        )
    smtp_cfg = get_smtp_cfg(db)
    if not smtp_cfg.get("host"):
        return RedirectResponse(
            "/admin/system-einstellungen?mail_error=SMTP-Host+nicht+konfiguriert", status_code=303
        )
    try:
        await send_test_mail(to=recipient, db=db)
        write_audit(db, "admin.system_settings.test_mail_sent", user_id=request.state.user.id)
        return RedirectResponse(
            f"/admin/system-einstellungen?mail_ok={recipient}", status_code=303
        )
    except MailConfigError as exc:
        return RedirectResponse(
            f"/admin/system-einstellungen?mail_error={str(exc)[:120]}", status_code=303
        )
    except Exception as exc:
        logger_admin.warning("Test-Mail fehlgeschlagen: %s", exc)
        return RedirectResponse(
            f"/admin/system-einstellungen?mail_error=Versand+fehlgeschlagen%3A+{str(exc)[:100]}",
            status_code=303,
        )


# ── Backup / Export ───────────────────────────────────────────────────────────

@router.get("/backup", response_class=HTMLResponse)
async def backup_page(request: Request, _=Depends(require_role("system_admin"))):
    return templates.TemplateResponse(request, "admin/backup.html", {
        "user": request.state.user,
    })


@router.get("/backup/json")
async def backup_json(request: Request, db: Session = Depends(get_db),
                      _=Depends(require_role("system_admin"))):
    import json as json_lib

    data = {
        "vehicles": [
            {"id": v.id, "dept_slug": v.dept.slug, "code": v.code, "name": v.name,
             "type": v.type, "is_first_train": v.is_first_train, "display_order": v.display_order}
            for v in db.query(VehicleMaster).all()
        ],
        "members": [
            {"id": m.id, "org_slug": m.org.slug if m.org else None,
             "lastname": m.lastname, "firstname": m.firstname, "phone": m.phone, "email": m.email,
             "qualifications": [{"code": mq.qualification.code} for mq in m.qualifications if mq.qualification]}
            for m in db.query(Member).all()
        ],
        "qualifications": [{"code": q.code, "label": q.label} for q in db.query(Qualification).all()],
        "alarm_types": [
            {"code": at.code, "category": at.category, "label": at.label,
             "default_first_train_only": at.default_first_train_only, "notify_neighbors": at.notify_neighbors}
            for at in db.query(AlarmType).all()
        ],
        "task_suggestions": [
            {
                "text": s.text,
                "alarms": [{"alarm_type_code": a.alarm_type_code, "display_order": a.display_order}
                           for a in sorted(s.alarm_assignments, key=lambda x: x.display_order)],
            }
            for s in db.query(TaskSuggestion).order_by(TaskSuggestion.id).all()
        ],
        "message_suggestions": [
            {
                "text": s.text,
                "alarms": [{"alarm_type_code": a.alarm_type_code, "display_order": a.display_order}
                           for a in sorted(s.alarm_assignments, key=lambda x: x.display_order)],
            }
            for s in db.query(MessageSuggestion).order_by(MessageSuggestion.id).all()
        ],
        "lage_hints": [{"text": h.text, "display_order": h.display_order} for h in db.query(LageHint).all()],
        "default_messages": [
            {
                "text": m.text,
                "alarms": [{"alarm_type_code": a.alarm_type_code,
                            "display_order": a.display_order, "due_after_sec": a.due_after_sec}
                           for a in sorted(m.alarm_assignments, key=lambda x: x.display_order)],
            }
            for m in db.query(DefaultMessage).order_by(DefaultMessage.id).all()
        ],
        "alarm_dispatch": [
            {"alarm_type_code": e.alarm_type_code, "vehicle_master_id": e.vehicle_master_id, "display_order": e.display_order}
            for e in db.query(AlarmDispatchVehicle).order_by(AlarmDispatchVehicle.alarm_type_code, AlarmDispatchVehicle.display_order).all()
        ],
    }
    from fastapi.responses import Response as FR
    content = json_lib.dumps(data, indent=2, ensure_ascii=False)
    return FR(
        content=content.encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=stammdaten-backup.json"},
    )


@router.post("/backup/restore", response_class=HTMLResponse)
async def backup_restore(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("system_admin")),
    backup_file: UploadFile = File(...),
):
    import json as json_lib
    lines = []
    try:
        raw = await backup_file.read()
        data = json_lib.loads(raw)

        dept_map = {d.slug: d for d in db.query(FireDept).all()}

        # Qualifications
        for q in data.get("qualifications", []):
            obj = db.query(Qualification).filter(Qualification.code == q["code"]).first()
            if not obj:
                db.add(Qualification(code=q["code"], label=q["label"]))
            else:
                obj.label = q["label"]
        db.flush()
        lines.append(f"Qualifikationen: {len(data.get('qualifications', []))} verarbeitet")

        # AlarmTypes
        for at in data.get("alarm_types", []):
            obj = db.get(AlarmType, at["code"])
            if not obj:
                db.add(AlarmType(**at))
            else:
                for k, v in at.items():
                    setattr(obj, k, v)
        db.flush()
        lines.append(f"Alarmtypen: {len(data.get('alarm_types', []))} verarbeitet")

        # Vehicles
        for v in data.get("vehicles", []):
            dept = dept_map.get(v.get("dept_slug"))
            if not dept:
                lines.append(f"  Warnung: Wehr '{v.get('dept_slug')}' nicht gefunden, Fahrzeug '{v.get('code')}' übersprungen")
                continue
            obj = db.query(VehicleMaster).filter(VehicleMaster.code == v["code"]).first()
            if not obj:
                db.add(VehicleMaster(
                    dept_id=dept.id, code=v["code"], name=v["name"],
                    type=v.get("type", ""), is_first_train=v.get("is_first_train", False),
                    display_order=v.get("display_order", 0),
                ))
            else:
                obj.name = v["name"]; obj.type = v.get("type", obj.type)
                obj.is_first_train = v.get("is_first_train", obj.is_first_train)
                obj.display_order = v.get("display_order", obj.display_order)
        db.flush()
        lines.append(f"Fahrzeuge: {len(data.get('vehicles', []))} verarbeitet")

        # TaskSuggestions – replace all
        if "task_suggestions" in data:
            db.query(TaskSuggestion).delete()
            db.flush()
            for item in data["task_suggestions"]:
                s = TaskSuggestion(text=item["text"])
                db.add(s); db.flush()
                for a in item.get("alarms", []):
                    db.add(TaskSuggestionAlarm(task_suggestion_id=s.id,
                                               alarm_type_code=a["alarm_type_code"],
                                               display_order=a.get("display_order", 0)))
            lines.append(f"Auftragsvorlagen: {len(data['task_suggestions'])} importiert")

        # MessageSuggestions – replace all
        if "message_suggestions" in data:
            db.query(MessageSuggestion).delete()
            db.flush()
            for item in data["message_suggestions"]:
                s = MessageSuggestion(text=item["text"])
                db.add(s); db.flush()
                for a in item.get("alarms", []):
                    db.add(MessageSuggestionAlarm(message_suggestion_id=s.id,
                                                  alarm_type_code=a["alarm_type_code"],
                                                  display_order=a.get("display_order", 0)))
            lines.append(f"Meldungsvorlagen: {len(data['message_suggestions'])} importiert")

        # LageHints – replace all
        if "lage_hints" in data:
            db.query(LageHint).delete()
            for h in data["lage_hints"]:
                db.add(LageHint(text=h["text"], display_order=h.get("display_order", 0)))
            lines.append(f"Lage-Hinweise: {len(data['lage_hints'])} importiert")

        # DefaultMessages – replace all
        if "default_messages" in data:
            db.query(DefaultMessage).delete()
            db.flush()
            for item in data["default_messages"]:
                m = DefaultMessage(text=item["text"])
                db.add(m); db.flush()
                for a in item.get("alarms", []):
                    db.add(DefaultMessageAlarm(default_message_id=m.id,
                                               alarm_type_code=a["alarm_type_code"],
                                               display_order=a.get("display_order", 0),
                                               due_after_sec=a.get("due_after_sec", 300)))
            lines.append(f"Default-Meldungen: {len(data['default_messages'])} importiert")

        # AlarmDispatch – replace all
        if "alarm_dispatch" in data:
            db.query(AlarmDispatchVehicle).delete()
            vm_map = {v.id: v for v in db.query(VehicleMaster).all()}
            for e in data["alarm_dispatch"]:
                vid = e.get("vehicle_master_id")
                if vid and vid in vm_map:
                    db.add(AlarmDispatchVehicle(
                        alarm_type_code=e["alarm_type_code"],
                        vehicle_master_id=vid,
                        display_order=e.get("display_order", 0),
                    ))
            lines.append(f"Ausrückordnung: {len(data['alarm_dispatch'])} Einträge importiert")

        db.commit()
        result = "Import erfolgreich:\n" + "\n".join(lines)
    except Exception as exc:
        db.rollback()
        result = f"Fehler beim Import: {exc}"

    return templates.TemplateResponse(request, "admin/backup.html", {
        "user": request.state.user,
        "import_result": result,
    })


# ── Lagekarte-Tokens ──────────────────────────────────────────────────────────

@router.get("/lagekarte-tokens", response_class=HTMLResponse)
async def lagekarte_tokens(
    request: Request, db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    from app.models.lagekarte import LagekarteToken
    from app.models.incident import Incident as _Incident
    from app.core.permissions import has_role
    user = request.state.user
    is_sysadmin = has_role(user, "system_admin")
    if is_sysadmin:
        tokens = db.query(LagekarteToken).order_by(LagekarteToken.created_at.desc()).all()
    else:
        tokens = (
            db.query(LagekarteToken)
            .filter(LagekarteToken.org_id == user.org_id)
            .order_by(LagekarteToken.created_at.desc())
            .all()
        )
    active_incidents = db.query(_Incident).filter(_Incident.status == "active").order_by(_Incident.id).all()
    return templates.TemplateResponse(request, "admin/lagekarte_tokens.html", {
        "user": user, "tokens": tokens, "new_token": None, "incidents": active_incidents,
    })


@router.post("/lagekarte-tokens/neu")
async def create_lagekarte_token(
    request: Request,
    label: str = Form(...),
    einsatz_id: str = Form(""),
    expires_at: str = Form(""),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    import hashlib, secrets as _sec
    from datetime import datetime as _dt
    from app.models.lagekarte import LagekarteToken
    from app.models.incident import Incident as _Incident

    user = request.state.user
    raw = "lkw_" + _sec.token_urlsafe(32)
    token_hash = hashlib.sha256(raw.encode()).hexdigest()

    expires = None
    if expires_at.strip():
        try:
            expires = _dt.fromisoformat(expires_at.strip())
        except ValueError:
            pass

    einsatz_id_int: int | None = None
    if einsatz_id.strip():
        try:
            einsatz_id_int = int(einsatz_id)
        except ValueError:
            pass

    tok = LagekarteToken(
        token_hash=token_hash,
        label=label,
        org_id=user.org_id,
        einsatz_id=einsatz_id_int,
        created_by_user_id=user.id,
        expires_at=expires,
    )
    db.add(tok)
    write_audit(db, "admin.lagekarte_token.created", user_id=user.id,
                payload={"label": label, "org_id": user.org_id})
    db.commit()

    from app.models.incident import Incident as _Incident2
    active_incidents = db.query(_Incident2).filter(_Incident2.status == "active").order_by(_Incident2.id).all()
    tokens = (
        db.query(LagekarteToken)
        .filter(LagekarteToken.org_id == user.org_id)
        .order_by(LagekarteToken.created_at.desc())
        .all()
    )
    return templates.TemplateResponse(request, "admin/lagekarte_tokens.html", {
        "user": user, "tokens": tokens, "new_token": raw, "incidents": active_incidents,
    })


@router.post("/lagekarte-tokens/{token_id}/sperren")
async def revoke_lagekarte_token(
    token_id: int, request: Request,
    db: Session = Depends(get_db), _=Depends(require_role("admin")),
):
    from app.models.lagekarte import LagekarteToken
    tok = db.get(LagekarteToken, token_id)
    if tok and not same_org_or_system_admin(request.state.user, tok.org_id):
        raise HTTPException(403, "Keine Berechtigung")
    if tok:
        tok.revoked_at = datetime.now(UTC)
        write_audit(db, "admin.lagekarte_token.revoked", user_id=request.state.user.id,
                    entity_type="lagekarte_token", entity_id=token_id)
        db.commit()
    return RedirectResponse("/admin/lagekarte-tokens", status_code=303)


# ── Push-Nachrichten ──────────────────────────────────────────────────────────

@router.get("/push-nachrichten", response_class=HTMLResponse)
async def push_notifications_page(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    user = request.state.user
    from app.services.push_service import _push_cfg
    from sqlalchemy.orm import joinedload
    cfg = _push_cfg(db)
    sub_count = db.query(PushSubscription).count()
    users_with_subs = (
        db.query(User)
        .join(PushSubscription, PushSubscription.user_id == User.id)
        .distinct()
        .order_by(User.display_name)
        .all()
    )
    push_logs = (
        db.query(PushLog)
        .options(joinedload(PushLog.target_user))
        .order_by(PushLog.sent_at.desc())
        .limit(30)
        .all()
    )
    return templates.TemplateResponse(request, "admin/push_notifications.html", {
        "user": user,
        "sub_count": sub_count,
        "users_with_subs": users_with_subs,
        "push_enabled": cfg["enabled"] and bool(cfg["private_key"]) and bool(cfg["public_key"]),
        "push_logs": push_logs,
        "sent": request.query_params.get("sent"),
        "error": request.query_params.get("error"),
    })


@router.post("/push-nachrichten/senden")
async def send_push_notification(
    request: Request,
    title: str = Form(...),
    body: str = Form(...),
    url: str = Form("/"),
    target: str = Form("all"),
    user_id: int | None = Form(None),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    from app.services.push_service import notify_all, notify_user
    title = title.strip()
    body = body.strip()
    url = url.strip() or "/"
    if not title or not body:
        return RedirectResponse("/admin/push-nachrichten?error=empty", status_code=303)
    if target == "user" and user_id:
        count = notify_user(db, user_id, title, body, url, source="admin_user")
    else:
        count = notify_all(db, title, body, url, source="admin_all")
    write_audit(db, "admin.push.manual_send", user_id=request.state.user.id,
                payload={"title": title, "target": target, "sent": count})
    db.commit()
    return RedirectResponse(f"/admin/push-nachrichten?sent={count}", status_code=303)


# ── Geräte-Login ──────────────────────────────────────────────────────────────

@router.get("/geraete-login", response_class=HTMLResponse)
async def device_tokens_list(
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    from sqlalchemy.orm import joinedload
    user = request.state.user
    tokens = (
        _org_filter(
            db.query(DeviceToken).join(User, DeviceToken.user_id == User.id),
            user, User.org_id,
        )
        .options(joinedload(DeviceToken.user))
        .order_by(DeviceToken.created_at.desc())
        .all()
    )
    roles = db.query(Role).all()
    all_orgs = db.query(FireDept).order_by(FireDept.name).all() if has_role(user, "system_admin") else []
    vehicles = (
        _org_filter(
            db.query(VehicleMaster).filter(VehicleMaster.active == True),  # noqa: E712
            user, VehicleMaster.dept_id,
        )
        .order_by(VehicleMaster.name)
        .all()
    )
    base_url = str(request.base_url).rstrip("/")
    return templates.TemplateResponse(request, "admin/device_tokens.html", {
        "user": user,
        "tokens": tokens,
        "roles": roles,
        "all_orgs": all_orgs,
        "vehicles": vehicles,
        "is_sysadmin": has_role(user, "system_admin"),
        "saved": request.query_params.get("saved"),
        "error": request.query_params.get("error"),
        "base_url": base_url,
        "new_token": None,
        "new_label": None,
        "new_qr_b64": None,
    })


@router.post("/geraete-login/neu")
async def create_device_token(
    request: Request,
    label: str = Form(...),
    role_codes: list[str] = Form([]),
    org_id: int | None = Form(None),
    vehicle_master_id: int | None = Form(None),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    import secrets as _secrets
    current_user = request.state.user
    label = label.strip()
    if not label:
        return RedirectResponse("/admin/geraete-login?error=empty_label", status_code=303)

    target_org_id = (
        org_id if has_role(current_user, "system_admin") and org_id and org_id != 0
        else current_user.org_id
    )

    # Rollen: system_admin darf nur system_admin vergeben
    is_sysadmin = has_role(current_user, "system_admin")
    allowed_codes = (
        {r.code for r in db.query(Role).all()}
        if is_sysadmin
        else {r.code for r in db.query(Role).all() if r.code != "system_admin"}
    )
    role_codes = [c for c in role_codes if c in allowed_codes]

    # Eindeutigen Username generieren
    raw_token = _secrets.token_urlsafe(32)
    username = f"geraet_{_secrets.token_hex(6)}"

    device_user = User(
        username=username,
        display_name=label,
        password_hash=hash_password(_secrets.token_urlsafe(32)),
        org_id=target_org_id,
        active=True,
        is_device=True,
    )
    db.add(device_user)
    db.flush()

    for code in role_codes:
        role = db.query(Role).filter(Role.code == code).first()
        if role:
            db.add(UserRole(user_id=device_user.id, role_id=role.id))

    token_hash = hash_api_key(raw_token)
    # Fahrzeug nur übernehmen wenn es in der eigenen Org liegt (oder system_admin)
    safe_vehicle_id: int | None = None
    if vehicle_master_id:
        vm = db.get(VehicleMaster, vehicle_master_id)
        if vm and (has_role(current_user, "system_admin") or vm.dept_id == target_org_id):
            safe_vehicle_id = vm.id
    dt = DeviceToken(label=label, token_hash=token_hash, user_id=device_user.id,
                     vehicle_master_id=safe_vehicle_id)
    db.add(dt)
    db.flush()

    write_audit(db, "admin.device_token.created", user_id=current_user.id,
                entity_type="device_token", entity_id=dt.id,
                payload={"label": label, "role_codes": role_codes})
    db.commit()

    # Token einmalig direkt in der Response zeigen — nie in URL/Logs
    from sqlalchemy.orm import joinedload as _jl
    all_tokens = (
        _org_filter(
            db.query(DeviceToken).join(User, DeviceToken.user_id == User.id),
            current_user, User.org_id,
        )
        .options(_jl(DeviceToken.user))
        .order_by(DeviceToken.created_at.desc())
        .all()
    )
    roles_all = db.query(Role).all()
    all_orgs = db.query(FireDept).order_by(FireDept.name).all() if has_role(current_user, "system_admin") else []
    vehicles_all = (
        _org_filter(
            db.query(VehicleMaster).filter(VehicleMaster.active == True),  # noqa: E712
            current_user, VehicleMaster.dept_id,
        )
        .order_by(VehicleMaster.name)
        .all()
    )
    base_url = str(request.base_url).rstrip("/")
    login_url = f"{base_url}/geraet-login?token={raw_token}"

    # QR-Code für den Device-Login generieren (gleiche Methode wie Einsatz-QR)
    new_qr_b64: str | None = None
    try:
        import base64
        import io
        import qrcode  # type: ignore
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=6, border=4)
        qr.add_data(login_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        new_qr_b64 = base64.b64encode(buf.getvalue()).decode()
    except Exception:
        pass  # QR-Generierung optional; URL bleibt weiterhin sichtbar

    return templates.TemplateResponse(request, "admin/device_tokens.html", {
        "user": current_user,
        "tokens": all_tokens,
        "roles": roles_all,
        "all_orgs": all_orgs,
        "vehicles": vehicles_all,
        "is_sysadmin": has_role(current_user, "system_admin"),
        "saved": "1",
        "error": None,
        "base_url": base_url,
        "new_token": raw_token,
        "new_label": label,
        "new_qr_b64": new_qr_b64,
    })


def _assert_device_token_access(dt: DeviceToken | None, current_user) -> None:
    """Raises 404 if not found, 403 if cross-org access by non-sysadmin."""
    if not dt:
        raise HTTPException(404)
    if not has_role(current_user, "system_admin"):
        owner_org = dt.user.org_id if dt.user else None
        if owner_org != current_user.org_id:
            raise HTTPException(404)  # 404 statt 403 – keine ID-Enumeration


@router.post("/geraete-login/{token_id}/widerrufen")
async def revoke_device_token(
    token_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    dt = db.get(DeviceToken, token_id)
    _assert_device_token_access(dt, request.state.user)
    dt.revoked_at = datetime.now(UTC)
    if dt.user:
        dt.user.active = False
    write_audit(db, "admin.device_token.revoked", user_id=request.state.user.id,
                entity_type="device_token", entity_id=token_id,
                payload={"label": dt.label})
    db.commit()
    return RedirectResponse("/admin/geraete-login?saved=1", status_code=303)


@router.post("/geraete-login/{token_id}/reaktivieren")
async def reactivate_device_token(
    token_id: int,
    request: Request,
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    dt = db.get(DeviceToken, token_id)
    _assert_device_token_access(dt, request.state.user)
    dt.revoked_at = None
    if dt.user:
        dt.user.active = True
    write_audit(db, "admin.device_token.reactivated", user_id=request.state.user.id,
                entity_type="device_token", entity_id=token_id,
                payload={"label": dt.label})
    db.commit()
    return RedirectResponse("/admin/geraete-login?saved=1", status_code=303)


@router.post("/geraete-login/{token_id}/fahrzeug")
async def assign_device_vehicle(
    token_id: int,
    request: Request,
    vehicle_master_id: int | None = Form(None),
    db: Session = Depends(get_db),
    _=Depends(require_role("admin")),
):
    dt = db.get(DeviceToken, token_id)
    _assert_device_token_access(dt, request.state.user)
    current_user = request.state.user
    safe_vehicle_id: int | None = None
    if vehicle_master_id:
        vm = db.get(VehicleMaster, vehicle_master_id)
        if vm and (has_role(current_user, "system_admin")
                   or vm.dept_id == (dt.user.org_id if dt.user else None)):
            safe_vehicle_id = vm.id
    dt.vehicle_master_id = safe_vehicle_id
    write_audit(db, "admin.device_token.vehicle_assigned", user_id=current_user.id,
                entity_type="device_token", entity_id=token_id,
                payload={"label": dt.label, "vehicle_master_id": safe_vehicle_id})
    db.commit()
    return RedirectResponse("/admin/geraete-login?saved=1", status_code=303)
